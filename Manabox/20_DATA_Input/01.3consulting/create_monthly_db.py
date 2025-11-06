#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
月次実績・予算データ統合スクリプト
#1（実績PL）と#2（予算PL）を月次で読み込み、マスターフォーマット#3に基づいて統合し、
SharePoint用のエクセルDB（月毎に縦に並ぶ形式）を作成
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def load_master_format(master_file):
    """マスターフォーマット#3を読み込む"""
    print(f"マスターフォーマットを読み込み中: {master_file}")
    df_master = pd.read_csv(master_file, encoding='utf-8-sig')
    
    # Account Codeを抽出（階層構造から）
    df_master['Account_Code'] = df_master['Account Code アカウント'].fillna('')
    
    # 空のAccount Codeは削除（サブアイテムは後で処理）
    df_master = df_master[df_master['Account_Code'] != '']
    
    # 必要な列のみ選択
    master_cols = {
        'Account_Code': 'Account Code アカウント',
        'Internal_code': 'Internal code',
        'Internal_Account_Name': 'Internal Account Name',
        'KEIEI_houkoku': 'KEIEI houkoku',
        '詳細': '詳細',
        'Description': 'Description',
        'Chỉ_tiêu': 'Chỉ tiêu',
        'Level': 'Level',
        'Source': 'Source'
    }
    
    df_master_clean = pd.DataFrame()
    for new_col, old_col in master_cols.items():
        if old_col in df_master.columns:
            df_master_clean[new_col] = df_master[old_col]
        else:
            df_master_clean[new_col] = ''
    
    return df_master_clean

def load_actual_data(file_path, sheet_name=None, account_code_col='Account Code', date_col='Date', amount_col='Amount'):
    """
    #1（実績）エクセルを読み込む
    
    Args:
        file_path: エクセルファイルのパス
        sheet_name: シート名（Noneの場合は最初のシート）
        account_code_col: Account Codeの列名
        date_col: 日付の列名
        amount_col: 金額の列名
    """
    print(f"実績データを読み込み中: {file_path}")
    
    try:
        # エクセルを読み込み
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(file_path)
        
        # 列名を標準化
        df.columns = df.columns.str.strip()
        
        # Account Code列を探す（複数の可能性）
        account_col = None
        for col in df.columns:
            if 'Account Code' in str(col) or 'アカウント' in str(col) or 'account' in str(col).lower():
                account_col = col
                break
        
        if account_col is None:
            raise ValueError(f"Account Code列が見つかりません。列名を確認してください: {df.columns.tolist()}")
        
        # 日付列を探す
        date_col_found = None
        for col in df.columns:
            if 'Date' in str(col) or '日付' in str(col) or '年月' in str(col) or '月' in str(col):
                date_col_found = col
                break
        
        if date_col_found is None:
            raise ValueError(f"日付列が見つかりません。列名を確認してください: {df.columns.tolist()}")
        
        # 金額列を探す（数値列を探す）
        amount_cols = []
        for col in df.columns:
            if col not in [account_col, date_col_found]:
                # 数値列かチェック
                if df[col].dtype in ['int64', 'float64']:
                    amount_cols.append(col)
        
        if not amount_cols:
            # すべての列を金額として扱う（Account Codeと日付以外）
            amount_cols = [col for col in df.columns if col not in [account_col, date_col_found]]
        
        # データをロング形式に変換
        df_long = pd.melt(
            df,
            id_vars=[account_col, date_col_found],
            value_vars=amount_cols,
            var_name='Account_Detail',
            value_name='Actual_Amount'
        )
        
        # 列名を標準化
        df_long = df_long.rename(columns={
            account_col: 'Account_Code',
            date_col_found: 'Date'
        })
        
        # Account Codeを文字列に変換
        df_long['Account_Code'] = df_long['Account_Code'].astype(str).str.strip()
        
        # 日付を変換
        df_long['Date'] = pd.to_datetime(df_long['Date'], errors='coerce')
        
        # 年月を抽出
        df_long['Year_Month'] = df_long['Date'].dt.to_period('M')
        df_long['Year_Month_Str'] = df_long['Year_Month'].astype(str)
        
        # 不要な行を削除
        df_long = df_long[df_long['Account_Code'].notna()]
        df_long = df_long[df_long['Account_Code'] != '']
        df_long = df_long[df_long['Account_Code'] != 'nan']
        
        print(f"実績データ: {len(df_long)}件のレコードを読み込みました")
        return df_long
        
    except Exception as e:
        print(f"エラー: {e}")
        raise

def load_budget_data(file_path, sheet_name=None, account_code_col='Account Code', date_col='Date', amount_col='Budget'):
    """
    #2（予算）エクセルを読み込む
    """
    print(f"予算データを読み込み中: {file_path}")
    
    try:
        # エクセルを読み込み
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(file_path)
        
        # 列名を標準化
        df.columns = df.columns.str.strip()
        
        # Account Code列を探す
        account_col = None
        for col in df.columns:
            if 'Account Code' in str(col) or 'アカウント' in str(col) or 'account' in str(col).lower():
                account_col = col
                break
        
        if account_col is None:
            raise ValueError(f"Account Code列が見つかりません。列名を確認してください: {df.columns.tolist()}")
        
        # 日付列を探す
        date_col_found = None
        for col in df.columns:
            if 'Date' in str(col) or '日付' in str(col) or '年月' in str(col) or '月' in str(col):
                date_col_found = col
                break
        
        if date_col_found is None:
            raise ValueError(f"日付列が見つかりません。列名を確認してください: {df.columns.tolist()}")
        
        # 金額列を探す
        amount_cols = []
        for col in df.columns:
            if col not in [account_col, date_col_found]:
                if df[col].dtype in ['int64', 'float64']:
                    amount_cols.append(col)
        
        if not amount_cols:
            amount_cols = [col for col in df.columns if col not in [account_col, date_col_found]]
        
        # データをロング形式に変換
        df_long = pd.melt(
            df,
            id_vars=[account_col, date_col_found],
            value_vars=amount_cols,
            var_name='Account_Detail',
            value_name='Budget_Amount'
        )
        
        # 列名を標準化
        df_long = df_long.rename(columns={
            account_col: 'Account_Code',
            date_col_found: 'Date'
        })
        
        # Account Codeを文字列に変換
        df_long['Account_Code'] = df_long['Account_Code'].astype(str).str.strip()
        
        # 日付を変換
        df_long['Date'] = pd.to_datetime(df_long['Date'], errors='coerce')
        
        # 年月を抽出
        df_long['Year_Month'] = df_long['Date'].dt.to_period('M')
        df_long['Year_Month_Str'] = df_long['Year_Month'].astype(str)
        
        # 不要な行を削除
        df_long = df_long[df_long['Account_Code'].notna()]
        df_long = df_long[df_long['Account_Code'] != '']
        df_long = df_long[df_long['Account_Code'] != 'nan']
        
        print(f"予算データ: {len(df_long)}件のレコードを読み込みました")
        return df_long
        
    except Exception as e:
        print(f"エラー: {e}")
        raise

def merge_data(master_df, actual_df, budget_df):
    """マスター、実績、予算を結合"""
    print("データを結合中...")
    
    # マスターと実績を結合
    merged = master_df.merge(
        actual_df,
        on='Account_Code',
        how='left',
        suffixes=('', '_actual')
    )
    
    # 予算を結合
    merged = merged.merge(
        budget_df,
        on=['Account_Code', 'Year_Month_Str'],
        how='left',
        suffixes=('', '_budget')
    )
    
    # 金額を0で埋める（NaNの場合）
    merged['Actual_Amount'] = merged['Actual_Amount'].fillna(0)
    merged['Budget_Amount'] = merged['Budget_Amount'].fillna(0)
    
    # 差異を計算
    merged['Variance'] = merged['Actual_Amount'] - merged['Budget_Amount']
    merged['Variance_Rate'] = merged.apply(
        lambda x: (x['Variance'] / x['Budget_Amount'] * 100) if x['Budget_Amount'] != 0 else 0,
        axis=1
    )
    
    # 列を並び替え
    column_order = [
        'Year_Month_Str',
        'Date',
        'Account_Code',
        'Internal_code',
        'Internal_Account_Name',
        'KEIEI_houkoku',
        '詳細',
        'Description',
        'Chỉ_tiêu',
        'Level',
        'Actual_Amount',
        'Budget_Amount',
        'Variance',
        'Variance_Rate',
        'Source'
    ]
    
    # 存在する列のみ選択
    available_cols = [col for col in column_order if col in merged.columns]
    merged = merged[available_cols]
    
    # 年月でソート
    merged = merged.sort_values(['Year_Month_Str', 'Account_Code'])
    
    print(f"結合完了: {len(merged)}件のレコード")
    return merged

def create_excel_db(merged_df, output_file):
    """SharePoint用のエクセルDBを作成（月毎に縦に並ぶ形式）"""
    print(f"エクセルDBを作成中: {output_file}")
    
    # エクセルライターを作成
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        merged_df.to_excel(writer, sheet_name='Data', index=False)
        
        # ワークブックとワークシートを取得
        workbook = writer.book
        worksheet = writer.sheets['Data']
        
        # スタイル設定
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ヘッダー行のスタイル
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # データ行のスタイル
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = border
                if cell.column == 1:  # Year_Month列
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.column >= 11:  # 金額列
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # 列幅の調整
        column_widths = {
            'A': 12,  # Year_Month_Str
            'B': 12,  # Date
            'C': 15,  # Account_Code
            'D': 15,  # Internal_code
            'E': 30,  # Internal_Account_Name
            'F': 20,  # KEIEI_houkoku
            'G': 25,  # 詳細
            'H': 30,  # Description
            'I': 30,  # Chỉ_tiêu
            'J': 8,   # Level
            'K': 18,  # Actual_Amount
            'L': 18,  # Budget_Amount
            'M': 18,  # Variance
            'N': 15,  # Variance_Rate
            'O': 10   # Source
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 行の高さ調整
        worksheet.row_dimensions[1].height = 25
        
        # フィルターを追加
        worksheet.auto_filter.ref = worksheet.dimensions
    
    print(f"エクセルDB作成完了: {output_file}")

def main():
    """メイン処理"""
    # ファイルパス（実際のファイルパスに置き換えてください）
    master_file = '/Users/sugenotomohiro/Desktop/CURSOR/Manabox/20_DATA_Input/01.3consulting/マスターフォーマット#3_最終版.csv'
    
    # 実績データと予算データのパス（実際のファイルパスに置き換えてください）
    actual_file = 'path/to/actual_data.xlsx'  # #1（実績）のエクセルファイル
    budget_file = 'path/to/budget_data.xlsx'  # #2（予算）のエクセルファイル
    
    output_file = '/Users/sugenotomohiro/Desktop/CURSOR/Manabox/20_DATA_Input/01.3consulting/月次DB_SharePoint用.xlsx'
    
    try:
        # マスターフォーマットを読み込み
        master_df = load_master_format(master_file)
        
        # 実績データを読み込み（ファイルが存在する場合のみ）
        if os.path.exists(actual_file):
            actual_df = load_actual_data(actual_file)
        else:
            print(f"警告: 実績データファイルが見つかりません: {actual_file}")
            print("サンプルデータを作成します...")
            # サンプルデータを作成
            actual_df = pd.DataFrame({
                'Account_Code': [],
                'Year_Month_Str': [],
                'Actual_Amount': []
            })
        
        # 予算データを読み込み（ファイルが存在する場合のみ）
        if os.path.exists(budget_file):
            budget_df = load_budget_data(budget_file)
        else:
            print(f"警告: 予算データファイルが見つかりません: {budget_file}")
            print("サンプルデータを作成します...")
            # サンプルデータを作成
            budget_df = pd.DataFrame({
                'Account_Code': [],
                'Year_Month_Str': [],
                'Budget_Amount': []
            })
        
        # データを結合
        merged_df = merge_data(master_df, actual_df, budget_df)
        
        # エクセルDBを作成
        create_excel_db(merged_df, output_file)
        
        print("\n処理完了！")
        print(f"出力ファイル: {output_file}")
        print(f"総レコード数: {len(merged_df)}件")
        print(f"年月数: {merged_df['Year_Month_Str'].nunique()}ヶ月")
        print(f"Account Code数: {merged_df['Account_Code'].nunique()}件")
        
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()

