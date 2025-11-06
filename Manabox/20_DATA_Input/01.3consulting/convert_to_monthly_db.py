#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
#3形式のCSVを月毎に縦に並ぶ形式のエクセルDBに変換
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import re

def parse_account_code_from_headers(csv_file):
    """CSVファイルからAccount Codeとメタ情報を抽出"""
    print("CSVファイルを解析中...")
    
    # CSVを読み込み（最初の8行をスキップせずに全て読み込む）
    with open(csv_file, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()
    
    # ヘッダー情報を抽出（行1-7）
    header_info = {
        'level1': lines[0].strip().split(',') if len(lines) > 0 else [],
        'level2': lines[1].strip().split(',') if len(lines) > 1 else [],
        'level3': lines[2].strip().split(',') if len(lines) > 2 else [],
        'internal_code': lines[3].strip().split(',') if len(lines) > 3 else [],
        'account_name': lines[4].strip().split(',') if len(lines) > 4 else [],
        'keiei_houkoku': lines[5].strip().split(',') if len(lines) > 5 else [],
        'detail_jp': lines[6].strip().split(',') if len(lines) > 6 else [],
        'description_en': lines[7].strip().split(',') if len(lines) > 7 else [],
    }
    
    # データ部分を読み込み（行8以降）
    data_df = pd.read_csv(csv_file, skiprows=7, encoding='utf-8-sig')
    
    return header_info, data_df

def extract_account_code_from_column(header_info, col_idx):
    """列インデックスからAccount Codeを抽出"""
    account_code = None
    level = 0
    
    # レベル3から順に確認
    if col_idx < len(header_info['level3']) and header_info['level3'][col_idx]:
        account_code = header_info['level3'][col_idx].strip()
        level = 3
    elif col_idx < len(header_info['level2']) and header_info['level2'][col_idx]:
        account_code = header_info['level2'][col_idx].strip()
        level = 2
    elif col_idx < len(header_info['level1']) and header_info['level1'][col_idx]:
        account_code = header_info['level1'][col_idx].strip()
        level = 1
    
    return account_code, level

def convert_to_long_format(header_info, data_df):
    """横形式のデータを縦形式（ロング形式）に変換"""
    print("データをロング形式に変換中...")
    
    result_rows = []
    
    # Date列を取得
    date_col = data_df.columns[0]  # 最初の列がDate
    
    # 各データ行を処理
    for row_idx, row in data_df.iterrows():
        date_value = row[date_col]
        
        # 日付を変換（Apr-25 -> 2025-04）
        date_str = convert_date_format(str(date_value))
        
        # 各列（Account Code）を処理
        for col_idx in range(1, len(data_df.columns)):  # Date列を除く
            col_name = data_df.columns[col_idx]
            amount_value = row[col_name]
            
            # Account Codeを抽出
            account_code, level = extract_account_code_from_column(header_info, col_idx)
            
            if account_code and account_code != '':
                # 数値を変換（カンマ区切りを削除）
                if pd.isna(amount_value) or amount_value == '-' or amount_value == '':
                    amount = 0
                else:
                    amount_str = str(amount_value).replace(',', '').replace('"', '')
                    try:
                        amount = float(amount_str)
                    except:
                        amount = 0
                
                # メタ情報を取得
                internal_code = header_info['internal_code'][col_idx] if col_idx < len(header_info['internal_code']) else ''
                account_name = header_info['account_name'][col_idx] if col_idx < len(header_info['account_name']) else ''
                keiei_houkoku = header_info['keiei_houkoku'][col_idx] if col_idx < len(header_info['keiei_houkoku']) else ''
                detail_jp = header_info['detail_jp'][col_idx] if col_idx < len(header_info['detail_jp']) else ''
                description_en = header_info['description_en'][col_idx] if col_idx < len(header_info['description_en']) else ''
                
                # 結果に追加
                result_rows.append({
                    'Year_Month': date_str,
                    'Date': date_value,
                    'Account_Code': account_code,
                    'Level': level,
                    'Internal_code': internal_code,
                    'Internal_Account_Name': account_name,
                    'KEIEI_houkoku': keiei_houkoku,
                    '詳細': detail_jp,
                    'Description': description_en,
                    'Amount': amount
                })
    
    # DataFrameに変換
    result_df = pd.DataFrame(result_rows)
    
    print(f"変換完了: {len(result_df)}件のレコード")
    return result_df

def convert_date_format(date_str):
    """日付形式を変換（Apr-25 -> 2025-04）"""
    # Apr-25形式を処理
    month_map = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
        'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
        'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
    }
    
    # Apr-25形式を検出
    match = re.match(r'([A-Za-z]+)-(\d+)', date_str)
    if match:
        month_str = match.group(1)
        year_str = match.group(2)
        
        if month_str in month_map:
            month = month_map[month_str]
            year = '20' + year_str if len(year_str) == 2 else year_str
            return f"{year}-{month}"
    
    # その他の形式はそのまま返す
    return date_str

def create_excel_db(df, output_file):
    """SharePoint用のエクセルDBを作成"""
    print(f"エクセルDBを作成中: {output_file}")
    
    # 列を並び替え
    column_order = [
        'Year_Month',
        'Date',
        'Account_Code',
        'Level',
        'Internal_code',
        'Internal_Account_Name',
        'KEIEI_houkoku',
        '詳細',
        'Description',
        'Amount'
    ]
    
    # 存在する列のみ選択
    available_cols = [col for col in column_order if col in df.columns]
    df_output = df[available_cols]
    
    # ソート
    df_output = df_output.sort_values(['Year_Month', 'Account_Code'])
    
    # エクセルライターを作成
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_output.to_excel(writer, sheet_name='Data', index=False)
        
        # ワークブックとワークシートを取得
        workbook = writer.book
        worksheet = writer.sheets['Data']
        
        # スタイル設定
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ヘッダー行のスタイル
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # データ行のスタイル
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = border
                if cell.column == 1:  # Year_Month列
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.column == 10:  # Amount列
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # 列幅の調整
        column_widths = {
            'A': 12,  # Year_Month
            'B': 12,  # Date
            'C': 15,  # Account_Code
            'D': 8,   # Level
            'E': 15,  # Internal_code
            'F': 30,  # Internal_Account_Name
            'G': 20,  # KEIEI_houkoku
            'H': 25,  # 詳細
            'I': 30,  # Description
            'J': 18   # Amount
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 行の高さ調整
        worksheet.row_dimensions[1].height = 25
        
        # フィルターを追加
        worksheet.auto_filter.ref = worksheet.dimensions
    
    print(f"エクセルDB作成完了: {output_file}")

def main():
    """メイン処理"""
    input_file = '/Users/sugenotomohiro/Desktop/CURSOR/Manabox/20_DATA_Input/01.3consulting/Glory　Format - #3 sample(DB for BI report).csv'
    output_file = '/Users/sugenotomohiro/Desktop/CURSOR/Manabox/20_DATA_Input/01.3consulting/月次DB_SharePoint用.xlsx'
    
    try:
        # CSVファイルを解析
        header_info, data_df = parse_account_code_from_headers(input_file)
        
        print(f"データ行数: {len(data_df)}行")
        print(f"列数: {len(data_df.columns)}列")
        
        # ロング形式に変換
        long_df = convert_to_long_format(header_info, data_df)
        
        # エクセルDBを作成
        create_excel_db(long_df, output_file)
        
        print("\n処理完了！")
        print(f"出力ファイル: {output_file}")
        print(f"総レコード数: {len(long_df)}件")
        print(f"年月数: {long_df['Year_Month'].nunique()}ヶ月")
        print(f"Account Code数: {long_df['Account_Code'].nunique()}件")
        
        # サンプルデータを表示
        print("\nサンプルデータ（最初の5件）:")
        print(long_df.head().to_string())
        
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()

